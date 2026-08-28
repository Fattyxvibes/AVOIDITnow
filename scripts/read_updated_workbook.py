#!/usr/bin/env python3
"""Parse the user-supplied workbook into validated JSON for the private importer."""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

EXPECTED_HEADERS = [
    "Category",
    "Listed brand",
    "Listed sub-product / offering",
    "Impact on source page",
    "Country shown",
    "Alternative 1 company",
    "Alternative 1 product/service",
    "Alternative 1 source",
    "Alternative 2 company",
    "Alternative 2 product/service",
    "Alternative 2 source",
    "Alternative 3 company",
    "Alternative 3 product/service",
    "Alternative 3 source",
    "Notes",
]


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: read_updated_workbook.py <xlsx>")
    workbook_path = Path(sys.argv[1]).resolve()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Alternatives" not in workbook.sheetnames or "Summary" not in workbook.sheetnames:
        raise SystemExit("Workbook must contain Summary and Alternatives sheets")

    sheet = workbook["Alternatives"]
    header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    headers = [text(value) for value in header_row[:len(EXPECTED_HEADERS)]]
    if headers != EXPECTED_HEADERS:
        raise SystemExit(f"Unexpected Alternatives headers: {headers!r}")

    summary = {text(row[0]): text(row[1]) for row in workbook["Summary"].iter_rows(values_only=True) if text(row[0])}
    capture_date = summary.get("Capture date", "")
    reviewed_at = datetime.strptime(capture_date, "%d %b %Y").strftime("%Y-%m-%d")
    source_url = "https://boycott-israel.org/boycott.html"
    source_label = "Supplied boycott and alternatives database"

    records = []
    workbook_row = 3
    for values in sheet.iter_rows(min_row=3, values_only=True):
        row = [text(value) for value in values[:len(EXPECTED_HEADERS)]]
        if not any(row):
            continue
        if len(row) != len(EXPECTED_HEADERS):
            raise SystemExit(f"Row {workbook_row} has {len(row)} columns; expected {len(EXPECTED_HEADERS)}")
        category, brand, subproduct, impact, country = row[:5]
        if not category or not brand or not subproduct or not impact:
            raise SystemExit(f"Row {workbook_row} is missing a required listing field")
        alternatives = []
        for position, company_index in enumerate((5, 8, 11), start=1):
            company, product_service, alternative_source = row[company_index:company_index + 3]
            if not company or not product_service or not alternative_source:
                raise SystemExit(f"Row {workbook_row} alternative {position} is incomplete")
            alternatives.append({
                "position": position,
                "company": company,
                "productService": product_service,
                "sourceUrl": alternative_source,
            })
        records.append({
            "workbookRow": workbook_row,
            "category": category,
            "listedBrand": brand,
            "listedSubproduct": subproduct,
            "impactOnSource": impact,
            "countryShown": country or None,
            "notes": row[14] or None,
            "sourceUrl": source_url,
            "sourceLabel": source_label,
            "sourceReviewedAt": reviewed_at,
            "alternatives": alternatives,
        })
        workbook_row += 1

    if len(records) != int(summary.get("Rows / listed items", "0")):
        raise SystemExit(f"Parsed {len(records)} records but Summary declares {summary.get('Rows / listed items')}")
    if len({record["workbookRow"] for record in records}) != len(records):
        raise SystemExit("Workbook row keys are not unique")

    print(json.dumps({
        "sourceWorkbook": workbook_path.name,
        "sourceUrl": source_url,
        "sourceReviewedAt": reviewed_at,
        "records": records,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
